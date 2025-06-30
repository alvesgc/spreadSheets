from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.chart import BarChart, Reference
import os

class RelatorioController:
    def gerar_excel(self, dados, nome_arquivo="relatorio_financeiro.xlsx"):
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório Financeiro"

        # Cabeçalho com formatação
        cabecalho = ["Categoria", "Valor (R$)", "Observações"]
        ws.append(cabecalho)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F81BD")
        center_align = Alignment(horizontal="center")
        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col in range(1, len(cabecalho) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border_style

        ws.append(["Salário", dados.salario, "Fixo mensal"])
        ws.append(["Renda extra", dados.renda_extra, "Outras fontes"])

        linha_gastos_inicio = ws.max_row + 1

        for nome, valor in dados.gastos:
            ws.append([nome, valor, ""])

        linha_gastos_fim = ws.max_row
        ws.append(["Total de gastos", f"=SUM(B{linha_gastos_inicio}:B{linha_gastos_fim})", ""])
        linha_total_gastos = ws.max_row
        ws.append(["Saldo final", f"=B2+B3-B{linha_total_gastos}", ""])

        linha_saldo = ws.max_row

        for row in ws.iter_rows(min_row=2, max_row=linha_saldo, min_col=2, max_col=2):
            for cell in row:
                cell.number_format = '"R$" #,##0.00'

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[column].width = max_length + 2

        chart = BarChart()
        data = Reference(ws, min_col=2, min_row=2, max_row=linha_saldo)
        cats = Reference(ws, min_col=1, min_row=2, max_row=linha_saldo)
        chart.add_data(data, titles_from_data=False)
        chart.set_categories(cats)
        chart.title = "Resumo Financeiro"
        ws.add_chart(chart, "E2")

        wb.save(nome_arquivo)
        return os.path.abspath(nome_arquivo)
